import hashlib
import os
import shutil
import time
import json
import re
import urllib.error
import urllib.request
from datetime import UTC, datetime
from pathlib import Path
from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, Header, HTTPException, UploadFile
from pydantic import BaseModel
from sqlmodel import Session, select
from ..config import settings
from ..database import engine, get_session
from ..models_lol import DataSource, ImportBatch, ImportError, LolTeamAlias, SourceRun

router = APIRouter(prefix="/api")
SOURCES = {
    "leaguepedia_schedule": "Leaguepedia Schedule",
    "leaguepedia_statistics": "Leaguepedia Statistics",
    "oracles_elixir": "Oracle's Elixir",
    "riot_datadragon": "Riot Data Dragon",
    "manual_odds_csv": "Manual odds CSV",
    "external_odds_api": "External odds API",
}


class SourceConfiguration(BaseModel):
    base_url: str = ""
    api_key: str | None = None
    clear_api_key: bool = False
    enabled: bool = True


class CustomSourceConfiguration(SourceConfiguration):
    display_name: str


def _now(): return datetime.now(UTC)


def _admin(token: str | None = Header(default=None, alias="X-Admin-Token")):
    if not settings.admin_token or token != settings.admin_token:
        raise HTTPException(403, "Administrative authentication required")


def _source(session, code):
    row = session.exec(select(DataSource).where(DataSource.code == code)).first()
    if row:
        return row
    if code not in SOURCES:
        raise HTTPException(404, "Unknown source")
    row = DataSource(
        code=code,
        display_name=SOURCES[code],
        configured=(code == "oracles_elixir"),
        enabled=(code == "oracles_elixir"),
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return row


def _config(row: DataSource) -> dict:
    try:
        return json.loads(row.config_json or "{}")
    except json.JSONDecodeError:
        return {}


def _view(row):
    view = {key: getattr(row, key) for key in (
        "code", "display_name", "enabled", "configured", "status", "last_run_at",
        "last_success_at", "last_error", "last_duration_ms", "records_received",
        "records_inserted", "records_updated", "records_skipped", "coverage", "next_run_at",
    )}
    config = _config(row)
    view["base_url"] = config.get("base_url", "")
    view["api_key_configured"] = bool(config.get("api_key"))
    view["custom"] = row.code not in SOURCES
    return view


def _configuration_view(row: DataSource) -> dict:
    config = _config(row)
    return {
        "code": row.code,
        "display_name": row.display_name,
        "base_url": config.get("base_url", ""),
        "api_key_configured": bool(config.get("api_key")),
        "enabled": row.enabled,
        "configured": row.configured,
        "custom": row.code not in SOURCES,
    }


def _validate_url(url: str) -> str:
    value = url.strip()
    if not value:
        return ""
    if not re.match(r"^https?://", value, flags=re.I):
        raise HTTPException(422, "La URL debe iniciar con http:// o https://")
    return value


@router.get("/sources")
def sources(session: Session = Depends(get_session)):
    builtin = [_source(session, code) for code in SOURCES]
    custom = [
        row for row in session.exec(select(DataSource).order_by(DataSource.display_name)).all()
        if row.code not in SOURCES
    ]
    return {"sources": [_view(row) for row in builtin + custom]}


@router.get("/sources/detail/{code}")
def source(code: str, session: Session = Depends(get_session)):
    return _view(_source(session, code))


@router.get("/sources/{code}/configuration", dependencies=[Depends(_admin)])
def source_configuration(code: str, session: Session = Depends(get_session)):
    return _configuration_view(_source(session, code))


@router.put("/sources/{code}/configuration", dependencies=[Depends(_admin)])
def save_source_configuration(
    code: str,
    payload: SourceConfiguration,
    session: Session = Depends(get_session),
):
    row = _source(session, code)
    config = _config(row)
    config["base_url"] = _validate_url(payload.base_url)
    if payload.clear_api_key:
        config.pop("api_key", None)
    elif payload.api_key is not None and payload.api_key.strip():
        config["api_key"] = payload.api_key.strip()
    row.config_json = json.dumps(config, sort_keys=True)
    row.enabled = payload.enabled
    row.configured = bool(config.get("base_url")) or code == "oracles_elixir"
    row.status = "healthy" if row.configured and row.enabled else "degraded"
    row.last_error = None
    session.add(row)
    session.commit()
    session.refresh(row)
    return _configuration_view(row)


@router.post("/sources/custom", dependencies=[Depends(_admin)])
def create_custom_source(
    payload: CustomSourceConfiguration,
    session: Session = Depends(get_session),
):
    name = payload.display_name.strip()
    if not name:
        raise HTTPException(422, "El nombre es obligatorio")
    base = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")[:40] or "api"
    code = f"custom_api_{base}"
    suffix = 2
    while session.exec(select(DataSource).where(DataSource.code == code)).first():
        code = f"custom_api_{base}_{suffix}"
        suffix += 1
    row = DataSource(code=code, display_name=name)
    session.add(row)
    session.commit()
    return save_source_configuration(code, payload, session)


def _test_configured_source(row: DataSource) -> tuple[bool, str | None]:
    config = _config(row)
    url = config.get("base_url", "")
    if not url:
        return row.configured, None if row.configured else "Configure una URL de API antes de probar"
    headers = {"User-Agent": settings.leaguepedia_user_agent}
    if config.get("api_key"):
        headers["Authorization"] = f"Bearer {config['api_key']}"
    try:
        request = urllib.request.Request(url, headers=headers, method="HEAD")
        with urllib.request.urlopen(request, timeout=8) as response:
            if response.status < 400:
                return True, None
    except urllib.error.HTTPError as exc:
        if exc.code not in {403, 405}:
            return False, f"HTTP {exc.code}"
    except Exception as exc:
        return False, str(exc)
    try:
        request = urllib.request.Request(url, headers=headers, method="GET")
        with urllib.request.urlopen(request, timeout=8) as response:
            return response.status < 400, None if response.status < 400 else f"HTTP {response.status}"
    except Exception as exc:
        return False, str(exc)


@router.post("/sources/{code}/test", dependencies=[Depends(_admin)])
def test(code: str, session: Session = Depends(get_session)):
    row = _source(session, code)
    started = _now()
    ok, error = _test_configured_source(row)
    finished = _now()
    run = SourceRun(
        source_code=code,
        job="test",
        status="success" if ok else "failed",
        started_at=started,
        finished_at=finished,
        duration_ms=int((finished - started).total_seconds() * 1000),
        error_message=error,
    )
    row.status = "healthy" if ok else "degraded"
    row.last_run_at = finished
    row.last_success_at = finished if ok else row.last_success_at
    row.last_error = error
    session.add_all([row, run])
    session.commit()
    return {"status": row.status, "run_id": run.id, "reason": error}


@router.post("/sources/{code}/sync", dependencies=[Depends(_admin)])
def sync(code: str, session: Session = Depends(get_session)):
    row = _source(session, code)
    run = SourceRun(
        source_code=code,
        job="sync",
        status="failed",
        started_at=_now(),
        finished_at=_now(),
        error_message="Adapter sync is not configured",
    )
    row.status = "degraded"
    row.last_error = run.error_message
    row.last_run_at = _now()
    session.add_all([row, run])
    session.commit()
    return {"status": "degraded", "run_id": run.id, "reason": run.error_message}

@router.get("/sources/runs")
def runs(session:Session=Depends(get_session)): return {"runs":session.exec(select(SourceRun).order_by(SourceRun.id.desc()).limit(100)).all()}

@router.get("/sources/runs/{run_id}")
def run_status(run_id: int, session: Session = Depends(get_session)):
    run = session.get(SourceRun, run_id)
    if not run:
        raise HTTPException(404, "Ejecución no encontrada")
    return run
@router.get("/imports")
def imports(session:Session=Depends(get_session)): return {"imports":session.exec(select(ImportBatch).order_by(ImportBatch.id.desc()).limit(100)).all()}
def _batch_view(batch: ImportBatch, error: str | None = None):
    return {
        "batch_id": batch.id,
        "filename": batch.filename,
        "status": batch.status,
        "games": batch.games_inserted,
        "teams": batch.teams_inserted,
        "players": batch.players_inserted,
        "inserted": batch.rows_received,
        "source_code": batch.source_code,
        "created_at": batch.created_at,
        "completed_at": batch.completed_at,
        "error": error,
    }


@router.get("/imports/{batch_id}")
def import_status(batch_id: int, session: Session = Depends(get_session)):
    batch = session.get(ImportBatch, batch_id)
    if not batch:
        raise HTTPException(404, "Importación no encontrada")
    latest_error = session.exec(
        select(ImportError).where(ImportError.batch_id == batch_id).order_by(ImportError.id.desc())
    ).first()
    return _batch_view(batch, latest_error.reason if latest_error else None)


@router.get("/imports/{batch_id}/errors")
def import_errors(batch_id:int,session:Session=Depends(get_session)): return {"errors":session.exec(select(ImportError).where(ImportError.batch_id==batch_id)).all()}
@router.get("/aliases/unresolved")
def aliases(session: Session = Depends(get_session)):
    from ..services.lol_team_aliases import EXHIBITION_TEAMS
    return {"aliases": list(EXHIBITION_TEAMS)}


@router.post("/aliases/synchronize", dependencies=[Depends(_admin)])
def synchronize_aliases(session: Session = Depends(get_session)):
    from ..services.lol_team_aliases import synchronize_known_aliases
    return synchronize_known_aliases(session)
@router.post("/sources/oracles/upload",dependencies=[Depends(_admin)])
async def oracle_upload(file:UploadFile=File(...),session:Session=Depends(get_session)):
    if not file.filename or Path(file.filename).suffix.lower() not in {".csv",".zip"}: raise HTTPException(400,"Only CSV or ZIP files are accepted")
    payload=await file.read()
    if not payload or len(payload)>100*1024*1024: raise HTTPException(400,"El archivo debe ser mayor a 0 y menor o igual a 100 MB")
    checksum=hashlib.sha256(payload).hexdigest()
    existing=session.exec(select(ImportBatch).where(ImportBatch.sha256==checksum)).first()
    if existing:return {"batch_id":existing.id,"status":existing.status,"duplicate":True}
    folder=Path(settings.lol_history_import_dir)/"inbox";folder.mkdir(parents=True,exist_ok=True)
    target=folder/Path(file.filename).name;target.write_bytes(payload)
    batch=ImportBatch(source_code="oracles_elixir",filename=target.name,sha256=checksum,status="pending",rows_received=0);session.add(batch);session.commit();session.refresh(batch)
    return {"batch_id":batch.id,"status":"pending","sha256":checksum}
@router.post("/imports/preview", dependencies=[Depends(_admin)])
async def preview_import(file: UploadFile = File(...), type: str = Form(default="oracle")):
    """Validate metadata and return at most twenty rows without scanning the file."""
    from io import BytesIO, StringIO
    from itertools import islice
    import csv

    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xls"}:
        raise HTTPException(400, "Preview supports CSV, XLSX and XLS")
    size = file.size
    if size is None:
        await file.seek(0)
        size = len(await file.read())
        await file.seek(0)
    if not size or size > 100 * 1024 * 1024:
        raise HTTPException(400, "El archivo debe ser mayor a 0 y menor o igual a 100 MB")
    try:
        if suffix == ".csv":
            await file.seek(0)
            sample = await file.read(min(size, 2 * 1024 * 1024))
            reader = csv.reader(StringIO(sample.decode("utf-8-sig", errors="replace")))
            headers = [str(value or "") for value in next(reader, [])]
            rows = list(islice(reader, 20))
            sheets = []
        else:
            await file.seek(0)
            raw = await file.read()
            import openpyxl
            book = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
            sheets = book.sheetnames
            iterator = book.active.iter_rows(values_only=True)
            headers = [str(value or "") for value in next(iterator, [])]
            rows = list(islice(iterator, 20))
        if not headers:
            raise ValueError("No se detectó una fila de encabezados")
        return {"type": type, "filename": file.filename, "file_size": size,
                "sheets": sheets, "headers": headers,
                "rows": [[str(value or "") for value in row] for row in rows],
                "preview_rows": len(rows), "valid_rows": None,
                "mapping": {header.lower().replace(" ", ""): header for header in headers}}
    except Exception as exc:
        raise HTTPException(400, f"Unreadable file: {exc}") from exc


def _process_import_batch(batch_id: int, target_name: str, replace_existing: bool = False) -> None:
    """Run a potentially long import after the upload response has been sent."""
    target = Path(target_name)
    started = _now()
    with Session(engine) as session:
        batch = session.get(ImportBatch, batch_id)
        if not batch:
            return
        batch.status = "running"
        run = SourceRun(source_code="oracles_elixir", job="upload_import", status="running", started_at=started)
        session.add_all([batch, run])
        session.commit()
        session.refresh(run)
        run_id = run.id
        try:
            from ..services.imports.oracles_elixir_importer import _import_csv_file
            from ..services.series_builder import rebuild_series

            result = _import_csv_file(session, str(target), replace=replace_existing)
            if not result["games"]:
                raise ValueError("No se importaron mapas válidos")
            rebuild_series(session)
            processed = Path(settings.lol_history_import_dir) / "processed"
            processed.mkdir(parents=True, exist_ok=True)
            destination = processed / target.name
            destination.unlink(missing_ok=True)
            for older in processed.glob(f"*{batch.filename}"):
                if older != destination:
                    older.unlink(missing_ok=True)
            target.replace(destination)
            finished = _now()
            batch = session.get(ImportBatch, batch_id)
            if replace_existing:
                previous = session.exec(select(ImportBatch).where(
                    ImportBatch.source_code == "oracles_elixir",
                    ImportBatch.filename == batch.filename,
                    ImportBatch.id != batch.id,
                    ImportBatch.status == "success",
                )).all()
                for old_batch in previous:
                    old_batch.status = "superseded"
                    session.add(old_batch)
            batch.status = "success"
            batch.games_inserted = result["games"]
            batch.teams_inserted = result["teams"]
            batch.players_inserted = result["players"]
            batch.completed_at = finished
            run = session.get(SourceRun, run_id)
            run.status = "success"
            run.finished_at = finished
            run.duration_ms = int((finished - started).total_seconds() * 1000)
            run.records_inserted = sum(result.values())
            run.details_json = str(result)
            session.add_all([batch, run])
            session.commit()
        except Exception as exc:
            session.rollback()
            finished = _now()
            batch = session.get(ImportBatch, batch_id)
            if batch:
                batch.status = "failed"
                batch.completed_at = finished
            failed = Path(settings.lol_history_import_dir) / "failed"
            failed.mkdir(parents=True, exist_ok=True)
            if target.exists():
                destination = failed / target.name
                destination.unlink(missing_ok=True)
                shutil.move(str(target), str(destination))
            run = session.get(SourceRun, run_id)
            run.status = "failed"
            run.finished_at = finished
            run.duration_ms = int((finished - started).total_seconds() * 1000)
            run.error_message = str(exc)
            records = [ImportError(batch_id=batch_id, reason=str(exc)), run]
            if batch:
                records.insert(0, batch)
            session.add_all(records)
            session.commit()


def _latest_processed_files(session: Session) -> list[Path]:
    processed = Path(settings.lol_history_import_dir) / "processed"
    batches = session.exec(select(ImportBatch).where(
        ImportBatch.source_code == "oracles_elixir",
        ImportBatch.status.in_(["success", "superseded"]),
    ).order_by(ImportBatch.created_at.desc())).all()
    filenames = []
    for batch in batches:
        if batch.filename not in filenames:
            filenames.append(batch.filename)
    files = []
    for filename in filenames:
        candidates = [path for path in processed.glob(f"*{filename}") if path.is_file()]
        if candidates:
            files.append(max(candidates, key=lambda path: path.stat().st_mtime))
    return files


def _synchronize_history(run_id: int) -> None:
    started = _now()
    with Session(engine) as session:
        run = session.get(SourceRun, run_id)
        run.status = "running"
        run.started_at = started
        session.add(run)
        session.commit()
        try:
            from ..services.imports.oracles_elixir_importer import _import_csv_file
            from ..services.series_builder import rebuild_series
            from ..services.lol_team_aliases import synchronize_known_aliases

            synchronize_known_aliases(session)
            files = _latest_processed_files(session)
            if not files:
                raise ValueError("No existen archivos Oracle procesados para sincronizar")
            totals = {"games": 0, "teams": 0, "players": 0}
            for filepath in sorted(files):
                result = _import_csv_file(session, str(filepath), replace=True)
                for key in totals:
                    totals[key] += result[key]
            rebuild_series(session)
            finished = _now()
            run = session.get(SourceRun, run_id)
            run.status = "success"
            run.finished_at = finished
            run.duration_ms = int((finished - started).total_seconds() * 1000)
            run.records_received = totals["games"]
            run.records_inserted = sum(totals.values())
            run.details_json = str({"files": [path.name for path in files], **totals})
            source = _source(session, "oracles_elixir")
            source.status = "healthy"
            source.last_run_at = finished
            source.last_success_at = finished
            source.last_error = None
            source.records_received = totals["games"]
            source.records_inserted = sum(totals.values())
            source.coverage = "complete"
            session.add_all([run, source])
            session.commit()
        except Exception as exc:
            session.rollback()
            finished = _now()
            run = session.get(SourceRun, run_id)
            run.status = "failed"
            run.finished_at = finished
            run.duration_ms = int((finished - started).total_seconds() * 1000)
            run.error_message = str(exc)
            source = _source(session, "oracles_elixir")
            source.status = "failed"
            source.last_run_at = finished
            source.last_error = str(exc)
            session.add_all([run, source])
            session.commit()


@router.post("/sources/synchronize", dependencies=[Depends(_admin)])
def synchronize_sources(background_tasks: BackgroundTasks, session: Session = Depends(get_session)):
    active = session.exec(select(SourceRun).where(
        SourceRun.job == "full_synchronize",
        SourceRun.status.in_(["queued", "running"]),
    ).order_by(SourceRun.id.desc())).first()
    if active:
        return {"run_id": active.id, "status": active.status, "already_running": True}
    run = SourceRun(source_code="oracles_elixir", job="full_synchronize", status="queued", started_at=_now())
    session.add(run)
    session.commit()
    session.refresh(run)
    background_tasks.add_task(_synchronize_history, run.id)
    return {"run_id": run.id, "status": "queued"}


async def _execute_odds_import(file: UploadFile, session: Session):
    if Path(file.filename or "").suffix.lower() != ".csv":
        raise HTTPException(400, "La carga de odds admite archivos CSV")
    payload = await file.read()
    if not payload or len(payload) > 100 * 1024 * 1024:
        raise HTTPException(400, "El archivo debe ser mayor a 0 y menor o igual a 100 MB")
    digest = hashlib.sha256(payload).hexdigest()
    existing = session.exec(select(ImportBatch).where(ImportBatch.sha256 == digest)).first()
    if existing:
        response = _batch_view(existing)
        response.update({"duplicate": True, "type": "manual_odds"})
        return response

    root = Path(settings.lol_odds_import_dir)
    inbox = root / "inbox"
    processed = root / "processed"
    failed = root / "failed"
    for folder in (inbox, processed, failed):
        folder.mkdir(parents=True, exist_ok=True)
    safe = Path(file.filename).name
    target = inbox / f"{digest[:12]}-{safe}"
    target.write_bytes(payload)
    started = _now()
    batch = ImportBatch(
        source_code="manual_odds_csv", filename=safe, sha256=digest,
        status="running", rows_received=0,
    )
    run = SourceRun(source_code="manual_odds_csv", job="upload_import", status="running", started_at=started)
    session.add_all([batch, run])
    session.commit()
    session.refresh(batch)
    session.refresh(run)
    try:
        from ..services.lol_odds_importer import import_odds_csv

        result = import_odds_csv(session, str(target))
        if result["inserted"] <= 0:
            raise ValueError("No se importaron odds. Verifique match_key, team_name y decimal_odds")
        destination = processed / target.name
        destination.unlink(missing_ok=True)
        for older in processed.glob(f"*{safe}"):
            if older != destination:
                older.unlink(missing_ok=True)
        target.replace(destination)
        previous = session.exec(select(ImportBatch).where(
            ImportBatch.source_code == "manual_odds_csv",
            ImportBatch.filename == safe,
            ImportBatch.id != batch.id,
            ImportBatch.status == "success",
        )).all()
        for old_batch in previous:
            old_batch.status = "superseded"
            session.add(old_batch)
        finished = _now()
        batch.status = "success"
        batch.rows_received = result["inserted"]
        batch.completed_at = finished
        run.status = "success"
        run.finished_at = finished
        run.duration_ms = int((finished - started).total_seconds() * 1000)
        run.records_received = result["inserted"]
        run.records_inserted = result["inserted"]
        source = _source(session, "manual_odds_csv")
        source.configured = True
        source.enabled = True
        source.status = "healthy"
        source.last_run_at = finished
        source.last_success_at = finished
        source.records_received += result["inserted"]
        source.records_inserted += result["inserted"]
        source.coverage = "partial"
        session.add_all([batch, run, source])
        session.commit()
        return {
            "batch_id": batch.id, "status": "success", "type": "manual_odds",
            "inserted": result["inserted"], "filename": safe,
        }
    except Exception as exc:
        session.rollback()
        batch = session.get(ImportBatch, batch.id)
        run = session.get(SourceRun, run.id)
        finished = _now()
        batch.status = "failed"
        batch.completed_at = finished
        run.status = "failed"
        run.finished_at = finished
        run.error_message = str(exc)
        if target.exists():
            destination = failed / target.name
            destination.unlink(missing_ok=True)
            target.replace(destination)
        session.add_all([batch, run, ImportError(batch_id=batch.id, reason=str(exc))])
        session.commit()
        raise HTTPException(400, str(exc)) from exc


@router.post("/sources/odds/upload", dependencies=[Depends(_admin)])
async def odds_upload(file: UploadFile = File(...), session: Session = Depends(get_session)):
    return await _execute_odds_import(file, session)


@router.post("/imports/execute", dependencies=[Depends(_admin)])
async def execute_import(
    file: UploadFile = File(...),
    type: str = Form(default="oracle"),
    replace_existing: bool = Form(default=False),
    session: Session = Depends(get_session),
):
    if type == "manual_odds":
        return await _execute_odds_import(file, session)
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".csv" or type not in {"oracle", "manual_statistics"}:
        raise HTTPException(400, "Automatic import currently supports Oracle/manual CSV")
    declared_size = file.size
    if not declared_size or declared_size > 100 * 1024 * 1024:
        raise HTTPException(400, "El archivo debe ser mayor a 0 y menor o igual a 100 MB")

    uploads = Path(settings.lol_history_import_dir) / "inbox" / "uploads"
    uploads.mkdir(parents=True, exist_ok=True)
    safe = Path(file.filename).name
    temporary = uploads / f".{time.time_ns()}-{safe}.uploading"
    digest_builder = hashlib.sha256()
    size = 0
    try:
        with temporary.open("wb") as output:
            while chunk := await file.read(1024 * 1024):
                size += len(chunk)
                if size > 100 * 1024 * 1024:
                    raise HTTPException(400, "El archivo supera el límite de 100 MB")
                digest_builder.update(chunk)
                output.write(chunk)
        if not size:
            raise HTTPException(400, "El archivo está vacío")
        digest = digest_builder.hexdigest()
        same_name = session.exec(select(ImportBatch).where(
            ImportBatch.source_code == "oracles_elixir",
            ImportBatch.filename == safe,
            ImportBatch.status.in_(["success", "superseded"]),
        ).order_by(ImportBatch.created_at.desc())).first()
        force_replace = replace_existing or same_name is not None
        duplicate = session.exec(select(ImportBatch).where(ImportBatch.sha256 == digest)).first()
        target = uploads / f"{digest[:12]}-{safe}"
        if duplicate:
            if not force_replace:
                temporary.unlink(missing_ok=True)
                response = _batch_view(duplicate)
                response["duplicate"] = True
                return response
            target.unlink(missing_ok=True)
            temporary.replace(target)
            duplicate.filename = safe
            duplicate.status = "queued"
            duplicate.completed_at = None
            session.add(duplicate)
            session.commit()
            return {"batch_id": duplicate.id, "status": "queued", "filename": safe, "replace_existing": True, "duplicate": True}
        temporary.replace(target)
        batch = ImportBatch(
            source_code="oracles_elixir",
            filename=safe,
            sha256=digest,
            status="queued",
            rows_received=0,
        )
        session.add(batch)
        session.commit()
        session.refresh(batch)
        return {
            "batch_id": batch.id,
            "status": "queued",
            "filename": safe,
            "file_size": size,
            "replace_existing": force_replace,
            "message": "Archivo recibido. La importación se iniciará en breve y continuará aunque se reinicie la aplicación.",
        }
    except Exception:
        temporary.unlink(missing_ok=True)
        raise


@router.get("/sources/{code}")
def source(code: str, session: Session = Depends(get_session)): return _view(_source(session, code))
